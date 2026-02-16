import openpyxl

PATH = r"tracker\HRP_Tracker.xlsx"
SHEETS = [
    "Horse_Profile",
    "Meters_History",
    "Timed_Works_Log",
    "Accessories_Log",
    "Conformation_Traits",
    "Race_Results",
]

wb = openpyxl.load_workbook(PATH)
for s in SHEETS:
    ws = wb[s]
    total = max(ws.max_row - 1, 0)
    caros = 0
    for (v,) in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if str(v).strip().lower() == "caros_compass":
            caros += 1
    print(f"{s}: total_rows={total}  Caros_Compass_rows={caros}")
