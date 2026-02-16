"""Read HRP_Tracker.xlsx and dump key sheets to a file."""
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("openpyxl not installed")
    sys.exit(1)

TRACKER = Path(r"c:\hrp-ops\tracker\HRP_Tracker.xlsx")
OUT = Path(r"c:\hrp-ops\tmp_tracker_dump.txt")

wb = openpyxl.load_workbook(TRACKER, read_only=True, data_only=True)

lines = []
for name in wb.sheetnames:
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    lines.append(f"\n## {name}  ({len(rows)-1} data rows)")
    if not rows:
        continue
    headers = [str(h) if h else f"col{i}" for i, h in enumerate(rows[0])]
    lines.append("| " + " | ".join(headers) + " |")
    lines.append("| " + " | ".join(["---"] * len(headers)) + " |")
    for row in rows[1:51]:
        cells = [str(c)[:80] if c is not None else "" for c in row]
        while len(cells) < len(headers):
            cells.append("")
        lines.append("| " + " | ".join(cells) + " |")

wb.close()

OUT.write_text("\n".join(lines), encoding="utf-8")
print(f"Wrote {len(lines)} lines to {OUT}")
