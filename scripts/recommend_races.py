"""recommend_races.py — Match horses to specific upcoming races.

Inputs:
  - inputs/YYYY-MM-DD/stable_snapshot.json
  - inputs/export/raw/_global/race_calendar.html
  - inputs/export/raw/_global/results.html
  - tracker/HRP_Tracker.xlsx (Nominations sheet)

Outputs:
  - outputs/race_calendar_YYYY-MM-DD.json
  - outputs/upcoming_entries_YYYY-MM-DD.json
  - reports/Race_Opportunities.md
  - outputs/approval_queue.json
"""

import json
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from bs4 import BeautifulSoup

ROOT = Path(__file__).resolve().parents[1]
RAW_GLOBAL = ROOT / "inputs" / "export" / "raw" / "_global"
OUTPUTS = ROOT / "outputs"
REPORTS = ROOT / "reports"
TRACKER = ROOT / "tracker" / "HRP_Tracker.xlsx"
TODAY = date.today().isoformat()


# ── Helpers ─────────────────────────────────────────────────────

def norm(name: str) -> str:
    return re.sub(r"[^a-z0-9]", "", name.lower())


def parse_distance_furlongs(dist_text: str) -> Optional[float]:
    """Convert distance text like '5f', '6 1/2f', '1m', '1 1/16m' to furlongs."""
    dist_text = dist_text.strip().lower()
    m = re.match(r"(\d+)\s*(\d+/\d+)?\s*([fm])", dist_text)
    if not m:
        return None
    whole = int(m.group(1))
    frac = 0.0
    if m.group(2):
        num, den = m.group(2).split("/")
        frac = int(num) / int(den)
    val = whole + frac
    if m.group(3) == "m":
        val *= 8  # 1 mile = 8 furlongs
    return val


# ── Race Calendar Parser ───────────────────────────────────────

def parse_race_calendar() -> List[Dict[str, Any]]:
    """Parse race_calendar.html into structured race list."""
    cal_path = RAW_GLOBAL / "race_calendar.html"
    if not cal_path.exists():
        return []

    soup = BeautifulSoup(cal_path.read_text(encoding="utf-8", errors="replace"), "html.parser")
    text = soup.get_text("\n", strip=True)
    lines = text.split("\n")

    races: List[Dict[str, Any]] = []
    current_date = ""
    current_track = ""
    race_block: List[str] = []

    for line in lines:
        line = line.strip()
        if not line:
            continue

        # Match date: M/D/YYYY or MM/DD/YYYY
        date_m = re.match(r"^(\d{1,2}/\d{1,2}/\d{4})$", line)
        if date_m:
            current_date = date_m.group(1)
            continue

        # Match track abbreviation (3-4 uppercase letters at start of context)
        track_m = re.match(r"^([A-Z]{2,5})$", line)
        if track_m and len(line) <= 5:
            current_track = track_m.group(1)
            continue

        # Match race class/conditions lines
        # These typically contain: Clm, OClm, Mdn, MdSpWt, Alw, Stk, etc.
        class_keywords = ["Clm", "OClm", "Mdn", "MdSpWt", "Alw", "Stk", "Hcp", "WCl",
                          "Opt", "Stakes", "Handicap", "Maiden", "Claiming",
                          "Statebred", "Fillies", "N1X", "N2X", "N3X", "N2L", "N3L",
                          "year-old", "three-year", "four-year", "Two-year"]

        if any(kw.lower() in line.lower() for kw in class_keywords) and len(line) > 5:
            # This is a race conditions line — extract what we can
            race: Dict[str, Any] = {
                "conditions": line,
                "date": current_date,
                "raw_text": line,
            }

            # Extract distance from conditions (e.g., "5f", "6 1/2f", "1m")
            dist_m = re.search(r"(\d+\s*\d*/?\d*\s*[fm])\b", line)
            if dist_m:
                race["distance"] = dist_m.group(1).strip()
                race["distance_f"] = parse_distance_furlongs(dist_m.group(1))

            # Extract surface
            if "Turf" in line:
                race["surface"] = "Turf"
            elif "Dirt" in line:
                race["surface"] = "Dirt"

            # Extract purse
            purse_m = re.search(r"\$[\d,.]+", line)
            if purse_m:
                race["purse"] = purse_m.group(0)

            # Extract claiming price
            claim_m = re.search(r"Claiming Price\s*\$?([\d,.]+)", line, re.I)
            if claim_m:
                race["claiming_price"] = claim_m.group(1)

            # Try to extract track from surrounding context
            if current_track:
                race["track"] = current_track

            # Try to extract time
            time_m = re.search(r"(\d{1,2}:\d{2})\s*(AM|PM)?", line, re.I)
            if time_m:
                race["post_time"] = time_m.group(0)

            races.append(race)

    # Also scan for structured table data
    for table in soup.find_all("table"):
        rows = table.find_all("tr")
        if len(rows) < 3:
            continue
        # Look for header with relevant columns
        for j, row in enumerate(rows):
            cells = [c.get_text(" ", strip=True) for c in row.find_all(["th", "td"])]
            header_text = " ".join(cells).upper()
            if "RACE" in header_text and ("TRACK" in header_text or "DISTANCE" in header_text or "DATE" in header_text):
                headers = [c.upper().strip() for c in cells]
                # Parse subsequent rows using these headers
                for data_row in rows[j + 1:]:
                    dcells = [c.get_text(" ", strip=True) for c in data_row.find_all(["td"])]
                    if len(dcells) < 3 or not any(dcells):
                        continue
                    race = {"raw_text": " | ".join(dcells)}
                    for k, hdr in enumerate(headers):
                        if k >= len(dcells):
                            break
                        val = dcells[k].strip()
                        if not val:
                            continue
                        if "DATE" in hdr:
                            race["date"] = val
                        elif "TIME" in hdr:
                            race["post_time"] = val
                        elif "TRACK" in hdr:
                            race["track"] = val
                        elif "DIST" in hdr:
                            race["distance"] = val
                            race["distance_f"] = parse_distance_furlongs(val)
                        elif "SURF" in hdr:
                            race["surface"] = val
                        elif any(x in hdr for x in ["CLASS", "COND", "RACE"]):
                            race.setdefault("conditions", val)
                        elif "PURSE" in hdr or "FEE" in hdr:
                            race["purse"] = val
                        elif hdr in ("#", "##", "RACE#"):
                            race["race_num"] = val
                    if race.get("track") or race.get("conditions"):
                        races.append(race)
                break

    # Deduplicate by raw_text
    seen = set()
    unique: List[Dict[str, Any]] = []
    for r in races:
        key = r.get("raw_text", "")
        if key and key not in seen:
            seen.add(key)
            unique.append(r)

    return unique


# ── Entries / Nominations Parser ───────────────────────────────

def parse_entries_from_tracker() -> List[Dict[str, str]]:
    """Load current entries/nominations from tracker XLSX."""
    if not TRACKER.exists():
        return []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(TRACKER), read_only=True)
        entries: List[Dict[str, str]] = []
        if "Nominations" in wb.sheetnames:
            ws = wb["Nominations"]
            headers: list = []
            for row in ws.iter_rows(values_only=True):
                vals = [str(c).strip() if c else "" for c in row]
                if not headers:
                    headers = vals
                    continue
                if vals[0]:
                    entry = dict(zip(headers, vals))
                    entries.append({
                        "horse_name": entry.get("Horse", ""),
                        "date": entry.get("Race Date", entry.get("Date", "")),
                        "track": entry.get("Track", ""),
                        "race_num": entry.get("Race#", entry.get("Race", "")),
                        "class": entry.get("Class", entry.get("Conditions", "")),
                        "source": "tracker_nominations",
                    })
        wb.close()
        return entries
    except Exception:
        return []


def parse_entries_from_snapshot(snapshot: Dict) -> List[Dict[str, str]]:
    """Extract entries from snapshot nomination data."""
    entries: List[Dict[str, str]] = []
    for h in snapshot.get("horses", []):
        for n in h.get("nominations", []):
            entries.append({
                "horse_name": h["name"],
                "date": n.get("date", ""),
                "track": n.get("track", ""),
                "race_num": n.get("race", n.get("field", "")),
                "source": "snapshot_profile",
            })
    return entries


# ── Horse→Race Scoring ─────────────────────────────────────────

def score_race_for_horse(horse: Dict, race: Dict) -> Dict[str, Any]:
    """Score how well a race matches a horse. Higher = better fit."""
    score = 0.0
    reasons: List[str] = []
    risks: List[str] = []

    # 1. Readiness (condition + stamina)
    cond_str = horse.get("condition", "100%").replace("%", "")
    stam_str = horse.get("stamina", "100%").replace("%", "")
    cond = int(cond_str) if cond_str.isdigit() else 100
    stam = int(stam_str) if stam_str.isdigit() else 100

    if stam < 70:
        score -= 20
        risks.append(f"Low stamina ({stam}%)")
    elif stam < 85:
        score -= 5
        risks.append(f"Moderate stamina ({stam}%)")
    else:
        score += 10
        reasons.append(f"Good stamina ({stam}%)")

    if cond >= 98:
        score += 5
        reasons.append(f"Peak condition ({cond}%)")

    # 2. Consistency
    consist_str = horse.get("consistency", "0").replace("+", "")
    try:
        consist = int(consist_str)
    except ValueError:
        consist = 0
    if consist >= 4:
        score += 5
        reasons.append(f"High consistency (+{consist})")
    elif consist <= 1:
        score -= 3
        risks.append(f"Low consistency ({consist})")

    # 3. Distance match
    horse_dist = horse.get("distance_meter", "0")
    race_dist_f = race.get("distance_f")
    if horse_dist and str(horse_dist).isdigit() and race_dist_f:
        hdist = int(horse_dist)
        # distance_meter is a 0-100 scale, higher = prefers longer
        # Approximate: 0-30 = sprint (5-6f), 30-60 = mid (6-8f), 60-100 = route (8f+)
        if race_dist_f <= 6:  # sprint
            if hdist <= 40:
                score += 8
                reasons.append("Distance fits (sprinter)")
            else:
                score -= 3
                risks.append("May prefer longer")
        elif race_dist_f <= 8:  # middle
            if 20 <= hdist <= 70:
                score += 8
                reasons.append("Distance fits (middle)")
        else:  # route
            if hdist >= 50:
                score += 8
                reasons.append("Distance fits (router)")
            else:
                score -= 3
                risks.append("May prefer shorter")

    # 4. Track match
    horse_track = horse.get("track", "")
    race_track = race.get("track", "")
    if horse_track and race_track:
        if race_track.upper() in horse_track.upper():
            score += 10
            reasons.append(f"Home track ({race_track})")
        else:
            # Shipping cost
            score -= 2
            risks.append(f"Ship to {race_track}")

    # 5. Record bonus
    record = horse.get("record", {})
    starts = int(record.get("starts", 0))
    wins = int(record.get("wins", 0))
    if starts > 0 and wins > 0:
        win_pct = wins / starts
        if win_pct >= 0.20:
            score += 5
            reasons.append(f"Proven winner ({wins}W/{starts}S)")
    elif starts == 0:
        reasons.append("First-time starter")

    # 6. Class compatibility (basic keyword matching)
    conditions = race.get("conditions", "").lower()
    if "maiden" in conditions or "mdn" in conditions:
        if starts > 0 and wins > 0:
            score -= 10  # Winner can't run in maiden
            risks.append("Already a winner — ineligible for maiden")
        elif starts == 0:
            score += 5
            reasons.append("Maiden eligible (unraced)")

    return {
        "score": round(score, 1),
        "reasons": reasons,
        "risks": risks,
    }


# ── Main Pipeline ──────────────────────────────────────────────

def load_snapshot() -> Dict[str, Any]:
    snap_path = ROOT / "inputs" / TODAY / "stable_snapshot.json"
    if not snap_path.exists():
        dirs = sorted((ROOT / "inputs").glob("20*-*-*"), reverse=True)
        for d in dirs:
            sp = d / "stable_snapshot.json"
            if sp.exists():
                snap_path = sp
                break
    return json.loads(snap_path.read_text(encoding="utf-8"))


def main() -> None:
    OUTPUTS.mkdir(parents=True, exist_ok=True)
    REPORTS.mkdir(parents=True, exist_ok=True)

    # 1. Parse race calendar
    print("Parsing race calendar...")
    races = parse_race_calendar()
    cal_output = {
        "date": TODAY,
        "source": "race_calendar.html",
        "total_races": len(races),
        "races": races,
    }
    cal_path = OUTPUTS / f"race_calendar_{TODAY}.json"
    cal_path.write_text(json.dumps(cal_output, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"  {cal_path.name}: {len(races)} races")

    # 2. Parse entries/nominations
    print("Parsing entries/nominations...")
    snapshot = load_snapshot()
    tracker_entries = parse_entries_from_tracker()
    snap_entries = parse_entries_from_snapshot(snapshot)

    # Merge and deduplicate
    all_entries = tracker_entries + snap_entries
    seen_keys = set()
    unique_entries: List[Dict] = []
    for e in all_entries:
        key = norm(e.get("horse_name", "")) + e.get("date", "") + e.get("track", "")
        if key not in seen_keys:
            seen_keys.add(key)
            unique_entries.append(e)

    entries_output = {
        "date": TODAY,
        "total_entries": len(unique_entries),
        "entries": unique_entries,
    }
    entries_path = OUTPUTS / f"upcoming_entries_{TODAY}.json"
    entries_path.write_text(json.dumps(entries_output, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"  {entries_path.name}: {len(unique_entries)} entries")

    # Build entered-horse lookup — ONLY from tracker (reliable source)
    # Snapshot profile nominations are unreliable in daily mode
    entered_norms = {norm(e["horse_name"]) for e in tracker_entries if e.get("horse_name")}

    # 3. Score horses against races
    print("Matching horses to races...")
    horses = snapshot.get("horses", [])

    # Classify horses
    recommendations: List[Dict[str, Any]] = []
    for h in horses:
        stam_str = h.get("stamina", "100%").replace("%", "")
        stam = int(stam_str) if stam_str.isdigit() else 100
        cond_str = h.get("condition", "100%").replace("%", "")
        cond = int(cond_str) if cond_str.isdigit() else 100
        h_norm = norm(h["name"])

        already_entered = h_norm in entered_norms

        # Determine readiness
        if stam < 70:
            status = "REST"
        elif stam < 85:
            status = "WORK"
        elif already_entered:
            status = "ENTERED"
        else:
            status = "READY"

        # Score against all races
        race_scores: List[Dict] = []
        for race in races:
            result = score_race_for_horse(h, race)
            if result["score"] > 0:
                race_scores.append({
                    "race": race,
                    "score": result["score"],
                    "reasons": result["reasons"],
                    "risks": result["risks"],
                })

        # Top 3
        race_scores.sort(key=lambda x: x["score"], reverse=True)
        top3 = race_scores[:3]

        recommendations.append({
            "horse": h["name"],
            "status": status,
            "stamina": stam,
            "condition": cond,
            "already_entered": already_entered,
            "top_races": top3,
            "record": h.get("record", {}),
            "track": h.get("track", "?"),
        })

    # 4. Generate Race_Opportunities.md
    print("Generating Race_Opportunities.md...")
    lines = [
        "# 🏁 Race Opportunities",
        f"> **Generated:** {TODAY} | **Races Parsed:** {len(races)} | **Entries Found:** {len(unique_entries)}",
        "",
    ]

    # Already entered
    entered = [r for r in recommendations if r["status"] == "ENTERED"]
    if entered:
        lines.append("## ✅ Already Entered / Nominated")
        lines.append("| Horse | Track | Stam | Cond |")
        lines.append("|-------|-------|------|------|")
        for r in entered:
            lines.append(f"| {r['horse']} | {r['track']} | {r['stamina']}% | {r['condition']}% |")
        lines.append("")

    # Ready with race targets
    ready = [r for r in recommendations if r["status"] == "READY" and r["top_races"]]
    if ready:
        lines.append("## 🎯 Ready — Top Race Targets (Approval Required)")
        for r in sorted(ready, key=lambda x: -x["top_races"][0]["score"] if x["top_races"] else 0):
            rec = r.get("record", {})
            rec_str = f"{rec.get('wins', 0)}W/{rec.get('starts', 0)}S" if rec.get("starts") else "Unraced"
            lines.append(f"### {r['horse']} ({rec_str}) — {r['track']}, Stam {r['stamina']}%")
            lines.append("| # | Race | Score | Fit | Risks |")
            lines.append("|---|------|-------|-----|-------|")
            for i, tr in enumerate(r["top_races"], 1):
                race = tr["race"]
                race_desc_parts = []
                if race.get("date"):
                    race_desc_parts.append(race["date"])
                if race.get("track"):
                    race_desc_parts.append(race["track"])
                if race.get("distance"):
                    race_desc_parts.append(race["distance"])
                if race.get("surface"):
                    race_desc_parts.append(race["surface"])
                conds = race.get("conditions", "")[:50]
                if conds:
                    race_desc_parts.append(conds)
                race_desc = " · ".join(race_desc_parts)
                fit = "; ".join(tr["reasons"][:3])
                risk = "; ".join(tr["risks"][:2]) if tr["risks"] else "—"
                lines.append(f"| {i} | {race_desc} | {tr['score']} | {fit} | {risk} |")
            lines.append("")

    # Ready but no matching races
    ready_no_match = [r for r in recommendations if r["status"] == "READY" and not r["top_races"]]
    if ready_no_match:
        lines.append("## ❓ Ready — No Matching Races Found")
        lines.append("| Horse | Track | Stam | Cond | Notes |")
        lines.append("|-------|-------|------|------|-------|")
        for r in ready_no_match:
            lines.append(f"| {r['horse']} | {r['track']} | {r['stamina']}% | {r['condition']}% | Check race calendar manually |")
        lines.append("")

    # Work / Rest
    work = [r for r in recommendations if r["status"] == "WORK"]
    rest = [r for r in recommendations if r["status"] == "REST"]
    if work:
        lines.append("## 🏋️ Needs Work (Stamina 70-84%)")
        for r in work:
            lines.append(f"- **{r['horse']}** — Stam {r['stamina']}%, Cond {r['condition']}%")
        lines.append("")
    if rest:
        lines.append("## 🛏️ Rest Required (Stamina <70%)")
        for r in rest:
            lines.append(f"- **{r['horse']}** — Stam {r['stamina']}%, Cond {r['condition']}%")
        lines.append("")

    lines.append("---")
    lines.append(f"*Auto-generated by `recommend_races.py` on {TODAY}*")
    opp_text = "\n".join(lines) + "\n"
    (REPORTS / "Race_Opportunities.md").write_text(opp_text, encoding="utf-8")
    print(f"  Race_Opportunities.md: {len(opp_text)} chars")

    # 5. Generate approval_queue.json
    queue: List[Dict] = []
    for r in recommendations:
        if r["status"] == "REST":
            queue.append({
                "horse": r["horse"],
                "action": "rest",
                "reason": f"Stamina {r['stamina']}% — rest to recover",
                "approval_required": False,
                "timestamp": datetime.now().isoformat(),
            })
        elif r["status"] == "WORK":
            queue.append({
                "horse": r["horse"],
                "action": "work",
                "reason": f"Stamina {r['stamina']}% — light works",
                "approval_required": False,
                "timestamp": datetime.now().isoformat(),
            })
        elif r["status"] == "ENTERED":
            queue.append({
                "horse": r["horse"],
                "action": "review_entry",
                "reason": "Already entered — review jockey/instructions",
                "approval_required": True,
                "timestamp": datetime.now().isoformat(),
            })
        elif r["top_races"]:
            for tr in r["top_races"][:1]:  # Top pick only
                race = tr["race"]
                queue.append({
                    "horse": r["horse"],
                    "action": "enter_race",
                    "race_date": race.get("date", ""),
                    "race_track": race.get("track", ""),
                    "race_distance": race.get("distance", ""),
                    "race_conditions": race.get("conditions", "")[:60],
                    "score": tr["score"],
                    "reasons": tr["reasons"],
                    "risks": tr["risks"],
                    "approval_required": True,
                    "timestamp": datetime.now().isoformat(),
                })

    aq_path = OUTPUTS / "approval_queue.json"
    aq_path.write_text(json.dumps(queue, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"  approval_queue.json: {len(queue)} items ({sum(1 for q in queue if q.get('approval_required'))} need approval)")

    # Summary
    print(f"\nSummary:")
    counts = {}
    for r in recommendations:
        counts[r["status"]] = counts.get(r["status"], 0) + 1
    for k, v in sorted(counts.items()):
        print(f"  {k}: {v}")


if __name__ == "__main__":
    main()
