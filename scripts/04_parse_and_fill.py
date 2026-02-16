import argparse
import json
import re
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

import openpyxl
from bs4 import BeautifulSoup

ROOT = Path(__file__).resolve().parents[1]
RAW_ROOT = ROOT / "inputs" / "export" / "raw"
MANIFEST_PATH = ROOT / "inputs" / "export" / "export_manifest.json"
TRACKER_PATH = ROOT / "tracker" / "HRP_Tracker.xlsx"
SUMMARY_PATH = ROOT / "outputs" / "daily_reports" / "IMPORT_SUMMARY.md"

SHEETS: Dict[str, Dict[str, Sequence[str]]] = {
    "Horse_Profile": {
        "headers": ["horse_id", "horse_name", "age_sex", "sire", "dam", "owner", "breeder", "current_track", "meters", "height", "weight", "profile_date", "source_file", "notes"],
        "keys": ["horse_id", "profile_date", "height", "weight", "current_track"],
    },
    "Meters_History": {
        "headers": ["horse_id", "horse_name", "date", "event", "track", "condition_from", "condition_to", "condition_delta", "stamina_from", "stamina_to", "stamina_delta", "consistency_from", "consistency_to", "distance_from", "distance_to", "height", "weight_from", "weight_to", "weight_delta", "raw_row", "source_file"],
        "keys": ["horse_id", "date", "event", "track", "condition_from", "condition_to", "stamina_from", "stamina_to"],
    },
    "Timed_Works_Log": {
        "headers": ["horse_id", "horse_name", "date", "track", "surface", "distance", "splits", "final_time", "rider", "assigned_weight", "start_code", "pace_instr_code", "effort_code", "rank", "meters_join_status", "meter_event", "meter_condition_pre", "meter_condition_post", "meter_stamina_pre", "meter_stamina_post", "meter_consistency_pre", "meter_consistency_post", "raw_row", "source_file"],
        "keys": ["horse_id", "date", "track", "distance", "surface", "final_time", "splits"],
    },
    "Accessories_Log": {
        "headers": ["horse_id", "horse_name", "date", "track", "blinkers_applied", "blinkers_apply", "blinkers_remove", "shadow_roll_applied", "shadow_roll_apply", "shadow_roll_remove", "lasix_applied", "lasix_apply", "lasix_remove", "bute_applied", "bute_apply", "bute_remove", "gelded_applied", "gelded_apply", "gelded_remove", "source_file", "notes"],
        "keys": ["horse_id", "date"],
    },
    "Conformation_Traits": {
        "headers": ["horse_id", "horse_name", "date", "height", "weight", "lumbosacral", "stifles", "rear_triangle", "back_leg_soundness", "humerus", "humerus_angle", "front_leg_soundness", "forehand", "source_file", "notes"],
        "keys": ["horse_id", "date"],
    },
    "Race_Results": {
        "headers": ["horse_id", "horse_name", "date", "track", "race_no", "distance", "surface", "finish", "raw_row", "notes", "source_file"],
        "keys": ["horse_id", "date", "track", "race_no"],
    },
}


def read_text(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="ignore") if path.exists() else ""


def norm(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def parse_mmddyyyy(value: str) -> str:
    m = re.search(r"\b(\d{1,2})/(\d{1,2})/(\d{4})\b", value)
    return f"{int(m.group(3)):04d}-{int(m.group(1)):02d}-{int(m.group(2)):02d}" if m else ""


def parse_ddmmmyy(value: str) -> str:
    m = re.match(r"^(\d{1,2})([A-Z]{3})(\d{2})$", value.upper())
    if not m:
        return ""
    mon = {"JAN": 1, "FEB": 2, "MAR": 3, "APR": 4, "MAY": 5, "JUN": 6, "JUL": 7, "AUG": 8, "SEP": 9, "OCT": 10, "NOV": 11, "DEC": 12}
    return f"{2000 + int(m.group(3)):04d}-{mon[m.group(2)]:02d}-{int(m.group(1)):02d}" if m.group(2) in mon else ""


def arrow(value: str) -> Tuple[str, str, str]:
    m = re.match(r"(.+?)\s*->\s*(.+?)(?:\s*\[([+-]?\d+)\])?$", norm(value))
    return (norm(m.group(1)), norm(m.group(2)), (m.group(3) or "").strip()) if m else ("", "", "")


def soup(html: str) -> Optional[BeautifulSoup]:
    return BeautifulSoup(html, "html.parser") if html else None


def file_soup(horse_dir: Path, name: str) -> Tuple[Optional[BeautifulSoup], str]:
    txt = read_text(horse_dir / name)
    return soup(txt), txt


def compact_cells(row) -> List[str]:
    out: List[str] = []
    for c in row.find_all(["th", "td"]):
        t = norm(c.get_text(" ", strip=True))
        if t:
            out.append(t)
    return out


def find_table(s: Optional[BeautifulSoup], required: Sequence[str]):
    if not s:
        return None
    req = [r.lower() for r in required]
    for t in s.find_all("table"):
        rs = t.find_all("tr")
        if not rs:
            continue
        cells = [norm(c.get_text(" ", strip=True)).lower() for c in rs[0].find_all(["th", "td"])]
        if all(any(r in c for c in cells) for r in req):
            return t
    return None


def extract_name(profile_s, works_s, meters_s) -> str:
    for s in [profile_s, works_s, meters_s]:
        if not s:
            continue
        for b in s.find_all("b"):
            t = norm(b.get_text(" ", strip=True))
            for p in ["Past Performance -", "Work Details -", "Meters -", "Conformation -", "Horse Accessories -"]:
                if t.startswith(p):
                    return norm(t.split("-", 1)[1])
    return ""


def parse_profile(horse_id: str, horse_name: str, text: str, source: str, warnings: List[str]) -> Dict[str, str]:
    t = norm(text)

    def g(p: str) -> str:
        m = re.search(p, t, flags=re.IGNORECASE)
        return norm(m.group(1)) if m else ""

    m = re.search(r"(\d+(?:\.\d+)?)h\s+(\d+)lbs", t, flags=re.IGNORECASE)
    row = {
        "horse_id": horse_id,
        "horse_name": horse_name,
        "age_sex": g(r"\b(Gr\.\s*[A-Za-z]\.\s*\d+\s*\([A-Za-z]{3}\))"),
        "sire": g(r"Sire:\s*([^:]+?)\s+Dam:"),
        "dam": g(r"Dam:\s*([^:]+?)\s+Br:"),
        "owner": g(rf"{re.escape(horse_name)}\s+Owner:\s*([^:]+?)\s+Condition:"),
        "breeder": g(r"Br:\s*([^:]+?)\s+WR:"),
        "current_track": g(r"Track:\s*([A-Za-z0-9]+)"),
        "meters": g(r"Distance:\s*([+-]?\d+)"),
        "height": m.group(1) if m else "",
        "weight": m.group(2) if m else "",
        "profile_date": parse_mmddyyyy(g(r"Prf Dt:\s*(\d{1,2}/\d{1,2}/\d{4})")),
        "source_file": source,
        "notes": "",
    }
    if not row["owner"] and not row["sire"] and not row["dam"]:
        row["notes"] = "Profile fields uncertain; page format may have changed."
        warnings.append(f"{horse_id}: sparse profile snapshot fields.")
    return row


def parse_meters(horse_id: str, horse_name: str, s: Optional[BeautifulSoup], source: str, warnings: List[str]) -> List[Dict[str, str]]:
    out: List[Dict[str, str]] = []
    t = find_table(s, ["date", "event", "track", "condition", "stamina", "consist", "dist", "weight"])
    if not t:
        warnings.append(f"{horse_id}: meters table not found.")
        return out
    for r in t.find_all("tr")[1:]:
        c = compact_cells(r)
        if len(c) < 9:
            continue
        cf, ct, cd = arrow(c[3])
        sf, st, sd = arrow(c[4])
        kf, kt, _ = arrow(c[5])
        df, dt, _ = arrow(c[6])
        wf, wt, wd = arrow(c[8])
        d = parse_mmddyyyy(c[0])
        if not d or not c[1]:
            continue
        out.append({
            "horse_id": horse_id,
            "horse_name": horse_name,
            "date": d,
            "event": c[1],
            "track": c[2],
            "condition_from": cf,
            "condition_to": ct,
            "condition_delta": cd,
            "stamina_from": sf,
            "stamina_to": st,
            "stamina_delta": sd,
            "consistency_from": kf,
            "consistency_to": kt,
            "distance_from": df,
            "distance_to": dt,
            "height": c[7],
            "weight_from": wf,
            "weight_to": wt,
            "weight_delta": wd,
            "raw_row": " | ".join(c),
            "source_file": source,
        })
    if not out:
        warnings.append(f"{horse_id}: no parsed meters rows.")
    return out


def parse_works(horse_id: str, horse_name: str, s: Optional[BeautifulSoup], source: str, warnings: List[str]) -> List[Dict[str, str]]:
    out: List[Dict[str, str]] = []
    if not s:
        warnings.append(f"{horse_id}: works_all.html missing.")
        return out
    txt = norm(" ".join(s.stripped_strings)).upper()
    m = re.search(r"\bWORKS:\s*(.*?)\s*(?:NOMINATIONS|ENTRIES|HORSE NOTES|$)", txt)
    if not m:
        warnings.append(f"{horse_id}: works segment not found.")
        return out
    chunks = [x.strip() for x in re.split(r"(?=\b\d{1,2}[A-Z]{3}\d{2}(?:-|\s))", m.group(1)) if x.strip()]
    for ch in chunks:
        dm = re.match(r"^(\d{1,2}[A-Z]{3}\d{2})(?:-|\s+)([A-Z0-9]+)\s*(.*)$", ch)
        if not dm:
            continue
        d = parse_ddmmmyy(dm.group(1))
        track = dm.group(2)
        rest = dm.group(3)
        toks = rest.split()
        dist, surf = "", ""
        for i, tok in enumerate(toks):
            if re.fullmatch(r"\d+F(?:\(T\))?", tok):
                dist = tok
                if i > 0 and re.fullmatch(r"[A-Z]{2,4}", toks[i - 1]):
                    surf = toks[i - 1]
                elif i + 1 < len(toks) and re.fullmatch(r"[A-Z]{1,4}", toks[i + 1]):
                    surf = toks[i + 1]
                break
        tp = re.findall(r"(\d{1,2}:\d{2}|:\d{2})\s+(\d+)", ch)
        splits = " | ".join(x for x, _ in tp[:-1]) if len(tp) > 1 else ""
        ftime = tp[-1][0] if tp else ""
        rank = tp[-1][1] if tp else ""
        rm = re.search(r"(?:\d{1,2}:\d{2}|:\d{2})\s+\d+\s+([A-Z][A-Z0-9'._-]+)\s+(\d{2,3})\b", ch)
        rider, wt = (rm.group(1), rm.group(2)) if rm else ("", "")
        sc, pc, ec = "", "", ""
        if wt and f" {wt} " in ch:
            tail = ch.split(f" {wt} ", 1)[1].split()
            code = [z for z in tail if re.fullmatch(r"[A-Z]{1,3}", z)]
            if len(code) >= 3:
                sc, pc, ec = code[0], code[1], code[2]
        if not d:
            continue
        out.append({
            "horse_id": horse_id,
            "horse_name": horse_name,
            "date": d,
            "track": track,
            "surface": surf,
            "distance": dist,
            "splits": splits,
            "final_time": ftime,
            "rider": rider,
            "assigned_weight": wt,
            "start_code": sc,
            "pace_instr_code": pc,
            "effort_code": ec,
            "rank": rank,
            "meters_join_status": "",
            "meter_event": "",
            "meter_condition_pre": "",
            "meter_condition_post": "",
            "meter_stamina_pre": "",
            "meter_stamina_post": "",
            "meter_consistency_pre": "",
            "meter_consistency_post": "",
            "raw_row": ch,
            "source_file": source,
        })
    if not out:
        warnings.append(f"{horse_id}: no parsed works rows.")
    return out


def parse_accessories(horse_id: str, horse_name: str, s: Optional[BeautifulSoup], txt: str, source: str, warnings: List[str]) -> List[Dict[str, str]]:
    if not s:
        warnings.append(f"{horse_id}: accessories.html missing.")
        return []
    t = find_table(s, ["accessory", "applied", "apply", "remove"])
    if not t:
        warnings.append(f"{horse_id}: accessories table not found.")
        return []
    d = parse_mmddyyyy(txt)
    tm = re.search(r"Track:\s*([A-Za-z0-9]+)", s.get_text(" ", strip=True), flags=re.IGNORECASE)
    row = {
        "horse_id": horse_id,
        "horse_name": horse_name,
        "date": d,
        "track": tm.group(1) if tm else "",
        "blinkers_applied": "", "blinkers_apply": "", "blinkers_remove": "",
        "shadow_roll_applied": "", "shadow_roll_apply": "", "shadow_roll_remove": "",
        "lasix_applied": "", "lasix_apply": "", "lasix_remove": "",
        "bute_applied": "", "bute_apply": "", "bute_remove": "",
        "gelded_applied": "", "gelded_apply": "", "gelded_remove": "",
        "source_file": source,
        "notes": "",
    }
    nm = {"BLINKERS": "blinkers", "SHADOW ROLL": "shadow_roll", "LASIX": "lasix", "BUTE": "bute", "GELDED": "gelded"}
    for tr in t.find_all("tr")[1:]:
        c = compact_cells(tr)
        if not c:
            continue
        n = c[0].upper()
        if n in nm:
            p = nm[n]
            row[f"{p}_applied"] = c[1] if len(c) > 1 else ""
            row[f"{p}_apply"] = c[2] if len(c) > 2 else ""
            row[f"{p}_remove"] = c[3] if len(c) > 3 else ""
    return [row]


def parse_conf(horse_id: str, horse_name: str, txt: str, source: str, warnings: List[str]) -> List[Dict[str, str]]:
    if not txt:
        warnings.append(f"{horse_id}: conformation.html missing.")
        return []
    t = norm(txt)

    def g(label: str) -> str:
        m = re.search(rf"{label}:\s*([^:]+?)(?:\s+[A-Za-z ]+:|$)", t, flags=re.IGNORECASE)
        return norm(m.group(1)) if m else ""

    row = {
        "horse_id": horse_id,
        "horse_name": horse_name,
        "date": parse_mmddyyyy(t),
        "height": g("Height"),
        "weight": g("Weight"),
        "lumbosacral": g("Lumbosacral"),
        "stifles": g("Stifles"),
        "rear_triangle": g("Rear Triangle"),
        "back_leg_soundness": g("Back Leg Soundness"),
        "humerus": g("Humerus"),
        "humerus_angle": g("Humerus Angle"),
        "front_leg_soundness": g("Front Leg Soundness"),
        "forehand": g("Forehand"),
        "source_file": source,
        "notes": "",
    }
    if not any([row["lumbosacral"], row["stifles"], row["rear_triangle"], row["back_leg_soundness"], row["humerus"], row["humerus_angle"], row["front_leg_soundness"], row["forehand"]]):
        row["notes"] = "Conformation traits uncertain; stored visible values only."
    return [row]


def parse_races(horse_id: str, horse_name: str, s: Optional[BeautifulSoup], txt: str, source: str, warnings: List[str]) -> List[Dict[str, str]]:
    out: List[Dict[str, str]] = []
    race_id_re = re.compile(r"^\d{1,2}[A-Za-z]{3}\d{2}-\d+[A-Z]{2,3}\b")
    race_id_parts_re = re.compile(r"^(\d{1,2}[A-Za-z]{3}\d{2})-(\d+)([A-Z]{2,3})$")
    dist_part_re = re.compile(r"^\d+(?:/\d+)?[A-Za-z]?$")

    def score_table(t) -> int:
        n = 0
        for tr in t.find_all("tr"):
            c = compact_cells(tr)
            if not c:
                continue
            for token in c[0].split():
                if race_id_re.match(token):
                    n += 1
        return n

    def parse_distance(tokens: List[str], start: int) -> str:
        if start >= len(tokens):
            return ""
        first = tokens[start]
        if not dist_part_re.match(first):
            return ""
        if start + 1 < len(tokens):
            second = tokens[start + 1]
            if re.fullmatch(r"\d+/\d+[A-Za-z]?", second):
                return f"{first} {second}"
            if re.fullmatch(r"[A-Za-z]{1,3}", second):
                return f"{first} {second}"
        return first

    chosen = None
    best = 0
    if s:
        for t in s.find_all("table"):
            sc = score_table(t)
            if sc > best:
                best = sc
                chosen = t

    if chosen:
        for tr in chosen.find_all("tr"):
            c = compact_cells(tr)
            if not c:
                continue
            line = c[0]
            if not race_id_re.match(line):
                continue
            if " " not in line:
                continue
            tokens = line.split()
            race_token = tokens[0]
            m = race_id_parts_re.match(race_token)
            if not m:
                continue
            date = parse_ddmmmyy(m.group(1))
            race_no = m.group(2)
            track = m.group(3).upper()
            surface = tokens[1] if len(tokens) > 1 else ""
            distance = parse_distance(tokens, 2)
            out.append({
                "horse_id": horse_id,
                "horse_name": horse_name,
                "race_token": race_token,
                "date": date,
                "track": track,
                "race_no": race_no,
                "distance": distance,
                "surface": surface,
                "finish": "",
                "raw_row": line,
                "notes": "",
                "source_file": source,
            })

    if out:
        return out

    # NEW: if there are no race-id tokens anywhere in the profile text, this horse is simply unraced.
    any_race_token = bool(re.search(r"\b\d{1,2}[A-Za-z]{3}\d{2}-\d+[A-Z]{2,3}\b", txt or ""))
    if not any_race_token:
        return []

    # Tokens exist but we couldn't parse combined-line rows => real parsing issue.
    warnings.append(f"{horse_id}: race rows not parsed; raw snippet stored.")
    return [{
        "horse_id": horse_id,
        "horse_name": horse_name,
        "race_token": "",
        "date": "",
        "track": "",
        "race_no": "",
        "distance": "",
        "surface": "",
        "finish": "",
        "raw_row": norm(txt)[:600],
        "notes": "Race parsing uncertain; stored raw profile snippet.",
        "source_file": source,
    }]


def join_works_meters(works: List[Dict[str, str]], meters: List[Dict[str, str]], warnings: List[str], horse_id: str) -> None:
    tw = [r for r in meters if r.get("event", "").upper() == "TIMED WORK"]
    bdt: Dict[Tuple[str, str], List[Dict[str, str]]] = {}
    bd: Dict[str, List[Dict[str, str]]] = {}
    for r in tw:
        bdt.setdefault((r.get("date", ""), r.get("track", "").upper()), []).append(r)
        bd.setdefault(r.get("date", ""), []).append(r)
    for w in works:
        cand = bdt.get((w.get("date", ""), w.get("track", "").upper()), [])
        chosen = None
        if len(cand) == 1:
            w["meters_join_status"] = "matched_exact"
            chosen = cand[0]
        elif len(cand) > 1:
            w["meters_join_status"] = "multiple_meter_matches"
        else:
            day = bd.get(w.get("date", ""), [])
            if len(day) == 1:
                w["meters_join_status"] = "matched_date_only"
                chosen = day[0]
            elif len(day) > 1:
                w["meters_join_status"] = "ambiguous_date_only"
            else:
                w["meters_join_status"] = "no_meter_match"
        if chosen:
            w["meter_event"] = chosen.get("event", "")
            w["meter_condition_pre"] = chosen.get("condition_from", "")
            w["meter_condition_post"] = chosen.get("condition_to", "")
            w["meter_stamina_pre"] = chosen.get("stamina_from", "")
            w["meter_stamina_post"] = chosen.get("stamina_to", "")
            w["meter_consistency_pre"] = chosen.get("consistency_from", "")
            w["meter_consistency_post"] = chosen.get("consistency_to", "")
        elif w["meters_join_status"] in {"multiple_meter_matches", "ambiguous_date_only"}:
            warnings.append(f"{horse_id}: uncertain work-meter join for {w.get('date')} {w.get('track')}.")


def parse_horse(horse_dir: Path, warnings: List[str]) -> Dict[str, List[Dict[str, str]]]:
    horse_id = horse_dir.name
    ps_all, ptxt_all = file_soup(horse_dir, "profile_allraces.html")
    ps_print, ptxt_print = file_soup(horse_dir, "profile_printable.html")

    ps = ps_all
    ptxt = ptxt_all
    pfile = "profile_allraces.html"
    if not ptxt:
        ps = ps_print
        ptxt = ptxt_print
        pfile = "profile_printable.html"

    ws, wtxt = file_soup(horse_dir, "works_all.html")
    ms, mtxt = file_soup(horse_dir, "meters.html")
    ass, atxt = file_soup(horse_dir, "accessories.html")
    _, ctxt = file_soup(horse_dir, "conformation.html")

    if not ptxt:
        warnings.append(f"{horse_id}: missing profile_allraces.html/profile_printable.html")
    if not wtxt:
        warnings.append(f"{horse_id}: missing works_all.html")
    if not mtxt:
        warnings.append(f"{horse_id}: missing meters.html")

    horse_name_parsed = extract_name(ps_print or ps_all, ws, ms)
    horse_name = horse_name_parsed or horse_dir.name.replace("_", " ").strip()
    if not horse_name_parsed:
        warnings.append(f"{horse_id}: horse name fallback from folder.")

    profile_text = norm(" ".join(ps.stripped_strings)) if ps else ptxt
    conformation_text = norm(" ".join(soup(ctxt).stripped_strings)) if ctxt else ""
    profile = parse_profile(horse_id, horse_name, profile_text, pfile, warnings)
    meters = parse_meters(horse_id, horse_name, ms, "meters.html", warnings)
    works = parse_works(horse_id, horse_name, ws, "works_all.html", warnings)
    join_works_meters(works, meters, warnings, horse_id)
    accessories = parse_accessories(horse_id, horse_name, ass, atxt, "accessories.html", warnings)
    conformation = parse_conf(horse_id, horse_name, conformation_text, "conformation.html", warnings)

    # NEW: avoid duplicate warnings when falling back from printable to allraces
    races: List[Dict[str, str]] = []

    attempt_warnings: List[str] = []
    if ptxt_print:
        races = parse_races(
            horse_id,
            horse_name,
            ps_print,
            norm(" ".join(ps_print.stripped_strings)) if ps_print else ptxt_print,
            "profile_printable.html",
            attempt_warnings,
        )

    if races and any(r.get("race_token", "") for r in races):
        warnings.extend(attempt_warnings)
    else:
        # discard first-attempt warnings to prevent duplicates
        races = []
        attempt_warnings = []
        if ptxt_all:
            races = parse_races(
                horse_id,
                horse_name,
                ps_all,
                norm(" ".join(ps_all.stripped_strings)) if ps_all else ptxt_all,
                "profile_allraces.html",
                attempt_warnings,
            )
        warnings.extend(attempt_warnings)

    return {
        "Horse_Profile": [profile],
        "Meters_History": meters,
        "Timed_Works_Log": works,
        "Accessories_Log": accessories,
        "Conformation_Traits": conformation,
        "Race_Results": races,
    }


def ensure_sheet(wb, sheet: str, headers: Sequence[str]):
    ws = wb[sheet] if sheet in wb.sheetnames else wb.create_sheet(sheet)
    first = [ws.cell(1, i + 1).value for i in range(len(headers))]
    if ws.max_row < 1 or all(v is None for v in first):
        for i, h in enumerate(headers, start=1):
            ws.cell(1, i, h)
    return ws


def header_idx(ws) -> Dict[str, int]:
    out: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v is not None:
            out[norm(str(v)).lower()] = c
    return out


def key_of(row: Dict[str, str], keys: Sequence[str]) -> str:
    return "|".join(norm(str(row.get(k, ""))).lower() for k in keys)


def existing(ws, keys: Sequence[str]) -> set:
    idx = header_idx(ws)
    seen = set()
    for r in range(2, ws.max_row + 1):
        vals: List[str] = []
        any_val = False
        for k in keys:
            c = idx.get(k.lower())
            val = ws.cell(r, c).value if c else ""
            s = norm(str(val)) if val is not None else ""
            vals.append(s.lower())
            any_val = any_val or bool(s)
        if any_val:
            seen.add("|".join(vals))
    return seen


def append_unique(ws, headers: Sequence[str], keys: Sequence[str], rows: List[Dict[str, str]]) -> int:
    seen = existing(ws, keys)
    n = 0
    for row in rows:
        k = key_of(row, keys)
        if k in seen:
            continue
        ws.append([row.get(h, "") for h in headers])
        seen.add(k)
        n += 1
    return n


def manifest_counts() -> Dict[str, int]:
    if not MANIFEST_PATH.exists():
        return {"horses_discovered": 0, "pages_exported": 0}
    d = json.loads(MANIFEST_PATH.read_text(encoding="utf-8"))
    return {"horses_discovered": int(d.get("horses_discovered", 0)), "pages_exported": int(d.get("pages_exported", 0))}


def write_summary(hparsed: int, parsed: Dict[str, int], appended: Dict[str, int], warnings: List[str]) -> None:
    mc = manifest_counts()
    SUMMARY_PATH.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        "# HRP Import Summary",
        "",
        f"- Horses discovered: {mc['horses_discovered']}",
        f"- Pages exported: {mc['pages_exported']}",
        f"- Horse folders parsed: {hparsed}",
        "",
        "## Parsed Rows",
    ]
    for s in SHEETS:
        lines.append(f"- {s}: {parsed.get(s, 0)}")
    lines += ["", "## Rows Appended"]
    for s in SHEETS:
        lines.append(f"- {s}: {appended.get(s, 0)}")
    lines += ["", "## Warnings"]
    lines += [f"- {w}" for w in warnings] if warnings else ["- None"]
    SUMMARY_PATH.write_text("\n".join(lines) + "\n", encoding="utf-8")


def discover(limit_horse: str = "") -> List[Path]:
    if not RAW_ROOT.exists():
        return []
    ds = sorted([p for p in RAW_ROOT.iterdir() if p.is_dir()])
    if not limit_horse:
        return ds
    return [d for d in ds if d.name.lower() == limit_horse.lower()]


def main() -> None:
    ap = argparse.ArgumentParser(description="Parse HRP exports into structured tracker truth tables")
    ap.add_argument("--dry-run", action="store_true", help="Parse only, do not modify workbook")
    ap.add_argument("--horse", default="", help="Single horse folder name, e.g. Caros_Compass")
    args = ap.parse_args()

    if not args.dry_run and not TRACKER_PATH.exists():
        raise FileNotFoundError(f"Tracker not found: {TRACKER_PATH}. Run scripts\\03_make_fresh_tracker.py first.")

    horse_dirs = discover(args.horse)
    warnings: List[str] = []
    rows: Dict[str, List[Dict[str, str]]] = {s: [] for s in SHEETS}
    for d in horse_dirs:
        parsed = parse_horse(d, warnings)
        for s, rs in parsed.items():
            rows[s].extend(rs)

    parsed_counts = {s: len(rows[s]) for s in SHEETS}
    appended_counts = {s: 0 for s in SHEETS}

    if not args.dry_run:
        wb = openpyxl.load_workbook(TRACKER_PATH)
        for s, cfg in SHEETS.items():
            ws = ensure_sheet(wb, s, cfg["headers"])
            appended_counts[s] = append_unique(ws, cfg["headers"], cfg["keys"], rows[s])
        wb.save(TRACKER_PATH)

    write_summary(len(horse_dirs), parsed_counts, appended_counts, warnings)

    print(f"Horse folders parsed: {len(horse_dirs)}")
    for s in SHEETS:
        print(f"{s}: parsed={parsed_counts.get(s, 0)} appended={appended_counts.get(s, 0)}")
    print(f"Warnings: {len(warnings)}")
    print(f"Summary written: {SUMMARY_PATH}")
    print("Dry run mode: workbook not modified." if args.dry_run else f"Workbook updated: {TRACKER_PATH}")


if __name__ == "__main__":
    main()
