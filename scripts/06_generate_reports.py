"""06_generate_reports.py — Generate markdown reports from stable_snapshot.json.

Produces:
  - reports/Stable_Dashboard.md (full rewrite)
  - reports/Weekly_Plan.md (full rewrite)
  - reports/Decisions_Log.md (append-only)
"""

import json
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List

ROOT = Path(__file__).resolve().parents[1]
REPORTS_DIR = ROOT / "reports"
TRACKER_PATH = ROOT / "tracker" / "HRP_Tracker.xlsx"


def load_tracker_nominations() -> List[Dict[str, str]]:
    """Load nominations from tracker XLSX Nominations sheet."""
    if not TRACKER_PATH.exists():
        return []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(TRACKER_PATH), read_only=True)
        noms: List[Dict[str, str]] = []
        if "Nominations" in wb.sheetnames:
            ws = wb["Nominations"]
            headers: list = []
            for row in ws.iter_rows(values_only=True):
                vals = [str(c).strip() if c else "" for c in row]
                if not headers:
                    headers = vals
                    continue
                if vals[0]:
                    noms.append(dict(zip(headers, vals)))
        wb.close()
        return noms
    except Exception:
        return []


def load_snapshot() -> Dict[str, Any]:
    """Load the latest stable_snapshot.json."""
    today = date.today().isoformat()
    path = ROOT / "inputs" / today / "stable_snapshot.json"
    if not path.exists():
        # Try yesterday
        from datetime import timedelta
        yesterday = (date.today() - timedelta(days=1)).isoformat()
        path = ROOT / "inputs" / yesterday / "stable_snapshot.json"
    if not path.exists():
        raise FileNotFoundError(
            f"No stable_snapshot.json found for {today} or previous day.\n"
            f"Run: python scripts/05_build_stable_snapshot.py"
        )
    return json.loads(path.read_text(encoding="utf-8"))


def generate_dashboard(snapshot: Dict[str, Any], tracker_noms: List[Dict[str, str]] = None) -> str:
    """Generate Stable_Dashboard.md."""
    d = snapshot.get("date", date.today().isoformat())
    balance = snapshot.get("balance", "?")
    horses = snapshot.get("horses", [])
    tracker_noms = tracker_noms or []

    lines = [
        f"# 🏇 Stable Dashboard",
        f"> **Generated:** {d} | **Balance:** ${balance} | **Horses:** {len(horses)}",
        "",
    ]

    # Build nomination lookup from tracker (source of truth)
    nom_by_horse: Dict[str, list] = {}
    for n in tracker_noms:
        horse = n.get("Horse", n.get("horse", ""))
        if horse:
            import re as _re
            norm = _re.sub(r"[^a-z0-9]", "", horse.lower())
            nom_by_horse.setdefault(norm, []).append(n)

    # Count horses with nominations (from tracker OR snapshot)
    with_noms = len(nom_by_horse)
    snap_noms = sum(1 for h in horses if h.get("nominations"))
    if snap_noms > with_noms:
        with_noms = snap_noms

    lines.append("## Quick Stats")
    lines.append(f"| Metric | Value |")
    lines.append(f"|--------|-------|")
    lines.append(f"| Active Horses | {len(horses)} |")
    lines.append(f"| Balance | ${balance} |")
    lines.append(f"| With Nominations | {with_noms} |")
    lines.append("")

    # Roster table
    lines.append("## Roster")
    lines.append("| Horse | Track | Cond | Stam | Consist | Mode | Noms |")
    lines.append("|-------|-------|------|------|---------|------|------|")

    for h in sorted(horses, key=lambda x: x.get("name", "")):
        name = h.get("name", "?")
        track = h.get("track", h.get("roster_track", "?"))
        cond = h.get("condition", h.get("roster_cond", "?"))
        stam = h.get("stamina", h.get("roster_stam", "?"))
        consist = h.get("consistency", h.get("roster_consist", "?"))
        mode_raw = h.get("roster_mode", "")
        mode = "🟢" if mode_raw and "R" not in mode_raw.upper() else "🔴" if mode_raw else "?"
        import re as _re
        h_norm = _re.sub(r"[^a-z0-9]", "", name.lower())
        snap_noms_count = len(h.get("nominations", []))
        tracker_noms_count = len(nom_by_horse.get(h_norm, []))
        noms = max(snap_noms_count, tracker_noms_count)
        nom_str = str(noms) if noms else "—"
        lines.append(f"| {name} | {track} | {cond} | {stam} | {consist} | {mode} | {nom_str} |")

    lines.append("")

    # Upcoming races — merge snapshot + tracker nominations
    lines.append("## Upcoming Races")
    if tracker_noms:
        lines.append("| Horse | Date | Track | Race# | Class |")
        lines.append("|-------|------|-------|-------|-------|")
        for n in tracker_noms:
            horse = n.get("Horse", n.get("horse", "?"))
            race_date = n.get("Race Date", n.get("Date", "?"))
            track = n.get("Track", "?")
            race_num = n.get("Race#", n.get("Race", "?"))
            cls = n.get("Class", n.get("Conditions", "?"))
            lines.append(f"| {horse} | {race_date} | {track} | {race_num} | {cls} |")
        lines.append("")
    else:
        nominated = [h for h in horses if h.get("nominations")]
        if nominated:
            lines.append("| Horse | Date | Track | Race |")
            lines.append("|-------|------|-------|------|")
            for h in nominated:
                for n in h["nominations"]:
                    lines.append(f"| {h['name']} | {n.get('date', '?')} | {n.get('track', '?')} | {n.get('race', '?')} |")
            lines.append("")
        else:
            lines.append("*No nominations found. Run weekly export for full data.*")
            lines.append("")

    # Horses needing attention
    def stam_pct(h: Dict) -> int:
        v = h.get("stamina", "100%").replace("%", "")
        return int(v) if v.isdigit() else 100
    low_stam = [h for h in horses if stam_pct(h) < 80]
    if low_stam:
        lines.append("## ⚠️ Low Stamina")
        for h in low_stam:
            lines.append(f"- **{h['name']}**: {h.get('stamina', '?')}")
        lines.append("")

    # Horse ability profiles (from horse_abilities.json)
    ab_path = ROOT / "outputs" / "horse_abilities.json"
    if ab_path.exists():
        abilities = json.loads(ab_path.read_text(encoding="utf-8"))
        if abilities:
            # Split into experienced (have speed) and unraced
            experienced = [a for a in abilities if a.get("best_speed", 0) > 0]
            experienced.sort(key=lambda x: -x.get("best_speed", 0))
            if experienced:
                lines.append("## 🎯 Horse Ability Profiles")
                lines.append("| Horse | Speed | Surface | Distance | Wet | Class | W% |")
                lines.append("|-------|:-----:|---------|----------|:---:|-------|:--:|")
                for a in experienced:
                    name = a["horse_name"]
                    spd = a.get("best_speed", 0)
                    spd_icon = "⚡" if spd >= 90 else ""
                    surf = a.get("preferred_surface", "?")
                    surf_icon = {"Turf": "🌿", "Dirt": "🏜️", "Both": "🔄"}.get(surf, "")
                    dist = a.get("preferred_distance", "?")
                    dist_icon = {"Sprint": "🏃", "Route": "🏇", "Both": "🔄"}.get(dist, "")
                    wet = a.get("wet_ability", 50)
                    wet_s = f"{'💧' if wet >= 85 else '⚠️' if wet < 70 else ''}{wet}"
                    cls = a.get("class_level", "?")
                    wr = a.get("win_rate", 0)
                    lines.append(f"| {name} | {spd_icon}{spd} | {surf_icon} {surf} | {dist_icon} {dist} | {wet_s} | {cls} | {wr:.0f}% |")
                lines.append("")

    # Report hub links
    lines.append("## 📊 Reports Hub")
    lines.append("| Report | Description |")
    lines.append("|--------|-------------|")
    lines.append("| [Race Opportunities](Race_Opportunities.md) | Model-scored race targets with field sizes |")
    lines.append("| [Training Plan](Training_Plan.md) | 14-day orders: Race/Work/Rest per horse |")
    lines.append("| [Approval Pack](Approval_Pack.md) | Checkbox approvals for entries |")
    lines.append("| [Trainer Scoreboard](Trainer_Scoreboard.md) | Win/Top3 rates, Brier score, benchmarks |")
    lines.append("| [Competitive Intel](Competitive_Intel.md) | Sitewide analysis vs top stables |")
    lines.append("")

    # Sitewide edge alerts (from patterns.json if available)
    patterns_path = ROOT / "outputs" / "sitewide" / "patterns.json"
    if patterns_path.exists():
        patterns = json.loads(patterns_path.read_text(encoding="utf-8"))
        rules = patterns.get("rules", [])
        high_conf = [r for r in rules if r.get("confidence") == "high"]
        if high_conf:
            lines.append("## 🔍 Sitewide Edge Alerts")
            for r in high_conf[:3]:
                lines.append(f"- {r['rule']}")
            lines.append("")

    # Peaking soon (from peak plan if available)
    import glob
    plan_files = sorted(glob.glob(str(ROOT / "outputs" / "peak_plan_*.json")), reverse=True)
    if plan_files:
        plan = json.loads(Path(plan_files[0]).read_text(encoding="utf-8"))
        peaking = plan.get("peaking_soon", [])
        at_risk = plan.get("at_risk", [])
        if peaking:
            lines.append("## ⚡ Peaking Soon")
            for name in peaking[:5]:
                lines.append(f"- **{name}** — race-ready, high readiness")
            lines.append("")
        if at_risk:
            lines.append("## 🚨 At Risk / Fatigued")
            for name in at_risk[:5]:
                lines.append(f"- **{name}** — needs rest or recovery")
            lines.append("")

    lines.append(f"---\n*Auto-generated by `06_generate_reports.py` on {d}*\n")
    return "\n".join(lines)


def generate_weekly_plan(snapshot: Dict[str, Any]) -> str:
    """Generate Weekly_Plan.md."""
    d = snapshot.get("date", date.today().isoformat())
    horses = snapshot.get("horses", [])
    balance = snapshot.get("balance", "?")

    lines = [
        f"# 📅 Weekly Plan",
        f"> **Week of:** {d} | **Balance:** ${balance}",
        "",
    ]

    # Race schedule from tracker nominations
    tracker_noms = load_tracker_nominations()
    if tracker_noms:
        lines.append("## Race Schedule")
        lines.append("| Date | Horse | Track | Race# | Class | Notes |")
        lines.append("|------|-------|-------|-------|-------|-------|")
        for n in tracker_noms:
            horse = n.get("Horse", n.get("horse", "?"))
            race_date = n.get("Race Date", n.get("Date", "?"))
            track = n.get("Track", "?")
            race_num = n.get("Race#", n.get("Race", "?"))
            cls = n.get("Class", n.get("Conditions", "?"))
            # Check stamina for this horse
            notes = ""
            for h in horses:
                import re as _re
                if _re.sub(r"[^a-z0-9]", "", h.get("name", "").lower()) == _re.sub(r"[^a-z0-9]", "", horse.lower()):
                    stam = h.get("stamina", "100%").replace("%", "")
                    if stam.isdigit() and int(stam) < 85:
                        notes = f"⚠️ Stam {stam}%"
                    break
            lines.append(f"| {race_date} | {horse} | {track} | {race_num} | {cls} | {notes} |")
        lines.append("")
        nom_count = len(set(n.get("Horse", n.get("horse", "")) for n in tracker_noms))
    else:
        nominated = [h for h in horses if h.get("nominations")]
        nom_count = len(nominated)
        if nominated:
            lines.append("## Race Schedule")
            lines.append("| Date | Horse | Track | Race |")
            lines.append("|------|-------|-------|------|")
            for h in nominated:
                for n in h["nominations"]:
                    lines.append(f"| {n.get('date', '?')} | {h['name']} | {n.get('track', '?')} | {n.get('race', '?')} |")
            lines.append("")

    # Training priorities
    no_noms = [h for h in horses if not h.get("nominations") and not h.get("record", {}).get("starts")]
    if no_noms:
        lines.append("## Training Priorities (Unraced, No Noms)")
        for h in no_noms:
            lines.append(f"- **{h['name']}** — Track: {h.get('track', '?')}, Stam: {h.get('stamina', '?')}")
        lines.append("")

    # Financial outlook
    lines.append("## Financial Outlook")
    lines.append(f"| Item | Value |")
    lines.append(f"|------|-------|")
    lines.append(f"| Current Balance | ${balance} |")
    lines.append(f"| Nominations Active | {nom_count} horses |")
    lines.append("")

    # Action items
    lines.append("## Action Items")
    lines.append("- [ ] Review race entries and set jockey instructions")
    lines.append("- [ ] Monitor stamina for nominated horses")
    lines.append("- [ ] Schedule works for unraced horses")
    lines.append("- [ ] Check for new race calendar entries")
    lines.append("")

    lines.append(f"---\n*Auto-generated by `06_generate_reports.py` on {d}*\n")
    return "\n".join(lines)


def append_decisions_log(snapshot: Dict[str, Any]) -> str:
    """Append a new session entry to Decisions_Log.md (never overwrite)."""
    d = snapshot.get("date", date.today().isoformat())
    log_path = REPORTS_DIR / "Decisions_Log.md"

    # Read existing content
    existing = ""
    if log_path.exists():
        existing = log_path.read_text(encoding="utf-8")

    # Don't add duplicate entry for same date
    if f"## Session: {d}" in existing:
        print(f"  Decisions_Log: skipped (entry for {d} already exists)")
        return existing

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
    balance = snapshot.get("balance", "?")
    horses = snapshot.get("horses", [])
    nominated = [h for h in horses if h.get("nominations")]

    entry_lines = [
        "",
        f"## Session: {d}",
        f"*Logged at {timestamp}*",
        "",
        f"- **Balance:** ${balance}",
        f"- **Active Horses:** {len(horses)}",
        f"- **Nominated:** {len(nominated)}",
        "",
        "### Pending Approvals",
        "*(none auto-generated — add manually)*",
        "",
        "---",
    ]

    new_entry = "\n".join(entry_lines)

    if not existing.strip():
        header = "# 📋 Decisions Log\n\n> Append-only log of decisions and approvals.\n\n---\n"
        return header + new_entry
    else:
        return existing.rstrip() + "\n" + new_entry


def main() -> None:
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    snapshot = load_snapshot()
    d = snapshot.get("date", "?")
    tracker_noms = load_tracker_nominations()
    print(f"  Tracker nominations loaded: {len(tracker_noms)} entries")

    # Dashboard (full rewrite)
    dashboard = generate_dashboard(snapshot, tracker_noms)
    (REPORTS_DIR / "Stable_Dashboard.md").write_text(dashboard, encoding="utf-8")
    print(f"  Stable_Dashboard.md — {len(dashboard)} chars")

    # Weekly Plan (full rewrite)
    plan = generate_weekly_plan(snapshot)
    (REPORTS_DIR / "Weekly_Plan.md").write_text(plan, encoding="utf-8")
    print(f"  Weekly_Plan.md — {len(plan)} chars")

    # Decisions Log (append-only)
    log = append_decisions_log(snapshot)
    (REPORTS_DIR / "Decisions_Log.md").write_text(log, encoding="utf-8")
    print(f"  Decisions_Log.md — {len(log)} chars")

    print(f"\nReports generated for {d}")


if __name__ == "__main__":
    main()
