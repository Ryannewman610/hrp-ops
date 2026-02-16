import shutil
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
TEMPLATE_PATH = ROOT / "tracker" / "TEMPLATE_HRP_Tracker.xlsx"
TRACKER_PATH = ROOT / "tracker" / "HRP_Tracker.xlsx"

LOG_SHEET_HINTS = (
    "log",
    "work",
    "race",
    "result",
    "pp",
    "history",
    "past",
)


def should_clear_sheet(title: str) -> bool:
    lower = title.lower()
    return any(hint in lower for hint in LOG_SHEET_HINTS)


def clear_data_rows(ws) -> int:
    if ws.max_row <= 1:
        return 0
    cleared = 0
    for row in range(2, ws.max_row + 1):
        row_had_value = False
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None and str(cell.value).strip() != "":
                row_had_value = True
            cell.value = None
        if row_had_value:
            cleared += 1
    return cleared


def main() -> None:
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(
            f"Template not found: {TEMPLATE_PATH}. Ensure filename is exactly TEMPLATE_HRP_Tracker.xlsx"
        )

    TRACKER_PATH.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(TEMPLATE_PATH, TRACKER_PATH)

    wb = openpyxl.load_workbook(TRACKER_PATH)

    total_cleared = 0
    selected = [ws for ws in wb.worksheets if should_clear_sheet(ws.title)]
    if not selected:
        selected = wb.worksheets

    for ws in selected:
        total_cleared += clear_data_rows(ws)

    wb.save(TRACKER_PATH)
    print(f"Created fresh tracker: {TRACKER_PATH}")
    print(f"Cleared data rows: {total_cleared}")


if __name__ == "__main__":
    main()
