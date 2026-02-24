"""00_data_integrity_audit.py — Cross-check all data sources for horse roster consistency.

Compares:
  1. Stable roster HTML (parsed horse names)
  2. Exported horse directories (inputs/export/raw/*)
  3. stable_snapshot.json
  4. Tracker Stable sheet (HRP_Tracker.xlsx)
  5. Nominations tracker sheet

Outputs a diff report and exits non-zero if mismatches found.
"""

import json
import re
import sys
from datetime import date
from pathlib import Path

from bs4 import BeautifulSoup

ROOT = Path(__file__).resolve().parents[1]
RAW_ROOT = ROOT / "inputs" / "export" / "raw"
GLOBAL_DIR = RAW_ROOT / "_global"
TRACKER_PATH = ROOT / "tracker" / "HRP_Tracker.xlsx"


def normalize(name: str) -> str:
    """Normalize horse name for comparison."""
    return re.sub(r"[^a-z0-9]", "", name.lower().strip())


def get_roster_horses() -> list[dict]:
    """Parse horse names from stable_roster.html."""
    roster_path = GLOBAL_DIR / "stable_roster.html"
    if not roster_path.exists():
        print(f"  WARN: {roster_path} not found")
        return []
    soup = BeautifulSoup(roster_path.read_text(encoding="utf-8", errors="replace"), "html.parser")

    horses = []
    # Find links to horse profiles
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if "/stables/" in href.lower() and "horse.aspx" in href.lower():
            name = a.get_text(strip=True)
            if name and len(name) > 1:
                horses.append({"name": name, "href": href, "norm": normalize(name)})

    # Deduplicate by normalized name
    seen = set()
    unique = []
    for h in horses:
        if h["norm"] not in seen:
            seen.add(h["norm"])
            unique.append(h)
    return unique


def get_exported_dirs() -> list[dict]:
    """Get list of exported horse directories."""
    if not RAW_ROOT.exists():
        return []
    dirs = []
    for d in sorted(RAW_ROOT.iterdir()):
        if d.is_dir() and d.name != "_global":
            name = d.name.replace("_", " ")
            dirs.append({"name": name, "dir": d.name, "norm": normalize(name)})
    return dirs


def get_snapshot_horses() -> list[dict]:
    """Get horse list from latest stable_snapshot.json."""
    today = date.today().isoformat()
    snap_path = ROOT / "inputs" / today / "stable_snapshot.json"
    if not snap_path.exists():
        dirs = sorted((ROOT / "inputs").glob("20*-*-*"), reverse=True)
        for d in dirs:
            sp = d / "stable_snapshot.json"
            if sp.exists():
                snap_path = sp
                break
    if not snap_path.exists():
        return []
    snap = json.loads(snap_path.read_text(encoding="utf-8"))
    return [{"name": h["name"], "norm": normalize(h["name"])} for h in snap.get("horses", [])]


def get_tracker_horses() -> list[dict]:
    """Get horse list from tracker Stable sheet."""
    if not TRACKER_PATH.exists():
        return []
    import openpyxl
    wb = openpyxl.load_workbook(str(TRACKER_PATH), read_only=True)
    horses = []
    # Try Stable sheet first, then Horse_Summary
    for sheet_name in ["Stable", "Horse_Summary"]:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0]:
                    name = str(row[0]).strip()
                    if name and name.lower() != "horse name":
                        horses.append({"name": name, "norm": normalize(name)})
            break
    wb.close()
    return horses


def get_tracker_nominations() -> list[dict]:
    """Get nominations from tracker Nominations sheet."""
    if not TRACKER_PATH.exists():
        return []
    import openpyxl
    wb = openpyxl.load_workbook(str(TRACKER_PATH), read_only=True)
    noms = []
    if "Nominations" in wb.sheetnames:
        ws = wb["Nominations"]
        headers = []
        for row in ws.iter_rows(values_only=True):
            vals = [str(c).strip() if c else "" for c in row]
            if not headers:
                headers = vals
                continue
            if vals[0]:
                entry = dict(zip(headers, vals))
                noms.append(entry)
    wb.close()
    return noms


def diff_sets(name_a: str, set_a: set, name_b: str, set_b: set) -> list[str]:
    """Compute differences between two name sets."""
    issues = []
    only_a = set_a - set_b
    only_b = set_b - set_a
    if only_a:
        issues.append(f"  In {name_a} but NOT in {name_b}: {sorted(only_a)}")
    if only_b:
        issues.append(f"  In {name_b} but NOT in {name_a}: {sorted(only_b)}")
    return issues


def main() -> None:
    print("=" * 60)
    print("DATA INTEGRITY AUDIT")
    print("=" * 60)

    # Gather all sources
    roster = get_roster_horses()
    exports = get_exported_dirs()
    snapshot = get_snapshot_horses()
    tracker = get_tracker_horses()
    noms = get_tracker_nominations()

    print(f"\n  Roster HTML:     {len(roster)} horses")
    print(f"  Exported dirs:   {len(exports)} directories")
    print(f"  Snapshot JSON:   {len(snapshot)} horses")
    print(f"  Tracker Stable:  {len(tracker)} horses")
    print(f"  Tracker Noms:    {len(noms)} entries")

    # Build normalized name sets
    roster_norms = {h["norm"] for h in roster}
    export_norms = {h["norm"] for h in exports}
    snap_norms = {h["norm"] for h in snapshot}
    tracker_norms = {h["norm"] for h in tracker}

    # Name lookup (norm -> display name)
    name_lookup: dict[str, str] = {}
    for src in [roster, exports, snapshot, tracker]:
        for h in src:
            if h["norm"] not in name_lookup:
                name_lookup[h["norm"]] = h["name"]

    issues: list[str] = []

    # Diff: Roster vs Exports
    print("\n--- Roster vs Exported Dirs ---")
    d = diff_sets("roster", roster_norms, "exports", export_norms)
    if d:
        issues.extend(d)
        for line in d:
            print(line)
    else:
        print("  MATCH ✓")

    # Diff: Roster vs Snapshot
    print("\n--- Roster vs Snapshot ---")
    d = diff_sets("roster", roster_norms, "snapshot", snap_norms)
    if d:
        issues.extend(d)
        for line in d:
            print(line)
    else:
        print("  MATCH ✓")

    # Diff: Roster vs Tracker
    print("\n--- Roster vs Tracker ---")
    d = diff_sets("roster", roster_norms, "tracker", tracker_norms)
    if d:
        issues.extend(d)
        for line in d:
            print(line)
    else:
        print("  MATCH ✓")

    # Diff: Exports vs Snapshot
    print("\n--- Exports vs Snapshot ---")
    d = diff_sets("exports", export_norms, "snapshot", snap_norms)
    if d:
        issues.extend(d)
        for line in d:
            print(line)
    else:
        print("  MATCH ✓")

    # Check for duplicates
    print("\n--- Duplicate Check ---")
    for src_name, src in [("roster", roster), ("exports", exports), ("snapshot", snapshot), ("tracker", tracker)]:
        norms = [h["norm"] for h in src]
        dupes = [n for n in set(norms) if norms.count(n) > 1]
        if dupes:
            dupe_names = [name_lookup.get(d, d) for d in dupes]
            issues.append(f"  Duplicates in {src_name}: {dupe_names}")
            print(f"  Duplicates in {src_name}: {dupe_names}")
    if not any("Duplicates" in i for i in issues):
        print("  No duplicates ✓")

    # Nominations info
    print(f"\n--- Nominations ---")
    print(f"  Tracker Nominations sheet: {len(noms)} entries")
    if noms:
        for n in noms[:5]:
            horse = n.get("Horse", n.get("horse", n.get(list(n.keys())[0], "?")))
            print(f"    {horse}: {n}")

    # Summary
    print("\n" + "=" * 60)
    if issues:
        print(f"AUDIT FOUND {len(issues)} ISSUE(S)")
        for i in issues:
            print(i)
        sys.exit(1)
    else:
        print("ALL DATA SOURCES AGREE ✓")
        sys.exit(0)


if __name__ == "__main__":
    main()
