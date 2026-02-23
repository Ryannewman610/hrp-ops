"""Extract latest meter reading per horse - fixed version."""
from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime

TRACKER = Path(r"c:\hrp-ops\tracker\HRP_Tracker.xlsx")
OUT = Path(r"c:\hrp-ops\tmp_latest_meters.txt")

wb = load_workbook(TRACKER, read_only=True, data_only=True)
ws = wb["Meters_History"]

rows = list(ws.iter_rows(min_row=1, values_only=True))
headers = list(rows[0])

# Print headers for debugging
print("Headers:", headers[:12])

# Use positional indexing based on the dump we already have:
# horse_id=0, horse_name=1, date=2, event=3, track=4, 
# cond_pre=5, cond_post=6, cond_delta=7, stam_pre=8, stam_post=9, stam_delta=10, 
# cons_pre=11, cons_post=12, cons_delta=13

latest = {}  # horse_name -> (date, event, track, cond_post, stam_post, cons_post)

for row in rows[1:]:
    if len(row) < 13:
        continue
    name = row[1]
    date_val = row[2]
    event = row[3]
    track = row[4]
    cond_post = row[6]
    stam_post = row[9]
    cons_post = row[12]
    
    if not name or not isinstance(name, str):
        continue
    
    # Skip junk/header rows - need at least a numeric condition
    try:
        cond_num = int(cond_post) if cond_post is not None else None
    except (ValueError, TypeError):
        continue
    
    if cond_num is None:
        continue
    
    # Convert date to comparable string
    if isinstance(date_val, datetime):
        ds = date_val.strftime("%Y-%m-%d")
    elif date_val:
        ds = str(date_val)
    else:
        ds = ""
    
    # Keep most recent
    if name not in latest or ds >= latest[name][0]:
        latest[name] = (ds, event, track, cond_post, stam_post, cons_post)

lines = []
lines.append("## Latest Meters (Feb 23 live scrape)")
lines.append("| Horse | Date | Event | Track | Cond | Stam | Cons |")
lines.append("| --- | --- | --- | --- | --- | --- | --- |")

for name in sorted(latest.keys()):
    d, ev, tr, c, s, co = latest[name]
    lines.append(f"| {name} | {d} | {ev} | {tr} | {c} | {s} | {co} |")

OUT.write_text("\n".join(lines), encoding="utf-8")
print(f"\nWrote {len(latest)} horses to {OUT}")
wb.close()
